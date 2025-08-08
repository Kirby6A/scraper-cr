"""
Data Export Service for exporting scraped data in multiple formats
"""
import csv
import json
import io
import os
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import tempfile
from pathlib import Path
from ..models import DataExport, TaskExecution, Task


class DataExportService:
    """Service for exporting data in various formats"""
    
    def __init__(self):
        self.export_dir = Path(tempfile.gettempdir()) / 'carbon_reform_exports'
        self.export_dir.mkdir(exist_ok=True)
    
    def export_data(
        self,
        export_id: str,
        format: str = 'json',
        config: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Export data based on DataExport configuration
        
        Args:
            export_id: DataExport ID
            format: Export format (csv, json, excel, parquet, xml)
            config: Format-specific configuration
        
        Returns:
            Export result with file path and metadata
        """
        try:
            export = DataExport.objects.get(id=export_id)
        except DataExport.DoesNotExist:
            return {'error': f'Export {export_id} not found'}
        
        # Update export status
        export.status = 'processing'
        export.save()
        
        try:
            # Get data to export
            data = self._get_export_data(export)
            
            # Export based on format
            if format == 'csv':
                result = self._export_csv(data, config or {})
            elif format == 'json':
                result = self._export_json(data, config or {})
            elif format == 'excel':
                result = self._export_excel(data, config or {})
            elif format == 'parquet':
                result = self._export_parquet(data, config or {})
            elif format == 'xml':
                result = self._export_xml(data, config or {})
            else:
                raise ValueError(f"Unsupported format: {format}")
            
            # Update export record
            export.status = 'completed'
            export.file_path = result['file_path']
            export.file_size = result['file_size']
            export.row_count = result['row_count']
            export.completed_at = datetime.now()
            export.save()
            
            return {
                'success': True,
                'export_id': str(export_id),
                'file_path': result['file_path'],
                'file_size': result['file_size'],
                'row_count': result['row_count'],
                'format': format
            }
        
        except Exception as e:
            # Update export with error
            export.status = 'failed'
            export.error_message = str(e)
            export.save()
            
            return {
                'success': False,
                'error': str(e),
                'export_id': str(export_id)
            }
    
    def export_task_data(
        self,
        task_id: str,
        format: str = 'json',
        filters: Optional[Dict[str, Any]] = None,
        config: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Export all execution data for a task
        
        Args:
            task_id: Task ID
            format: Export format
            filters: Filters to apply to executions
            config: Format-specific configuration
        
        Returns:
            Export result
        """
        # Get executions
        executions = TaskExecution.objects.filter(
            task_id=task_id,
            status='success'
        )
        
        # Apply filters if provided
        if filters:
            if 'date_from' in filters:
                executions = executions.filter(started_at__gte=filters['date_from'])
            if 'date_to' in filters:
                executions = executions.filter(started_at__lte=filters['date_to'])
        
        # Extract data
        data = []
        for execution in executions:
            if execution.scraped_data:
                if isinstance(execution.scraped_data, list):
                    data.extend(execution.scraped_data)
                else:
                    data.append(execution.scraped_data)
        
        # Create export record
        export = DataExport.objects.create(
            task_id=task_id,
            format=format,
            filters=filters or {},
            export_config=config or {},
            status='processing'
        )
        
        # Export data
        return self.export_data(str(export.id), format, config)
    
    def _get_export_data(self, export: DataExport) -> List[Dict]:
        """Get data to export based on export configuration"""
        data = []
        
        # Get data from executions
        if export.executions.exists():
            for execution in export.executions.all():
                if execution.scraped_data:
                    if isinstance(execution.scraped_data, list):
                        data.extend(execution.scraped_data)
                    else:
                        data.append(execution.scraped_data)
        
        # Or get data from task
        elif export.task:
            executions = TaskExecution.objects.filter(
                task=export.task,
                status='success'
            )
            
            # Apply filters
            if export.filters:
                # Apply date filters
                if 'date_from' in export.filters:
                    executions = executions.filter(
                        started_at__gte=export.filters['date_from']
                    )
                if 'date_to' in export.filters:
                    executions = executions.filter(
                        started_at__lte=export.filters['date_to']
                    )
            
            for execution in executions:
                if execution.scraped_data:
                    if isinstance(execution.scraped_data, list):
                        data.extend(execution.scraped_data)
                    else:
                        data.append(execution.scraped_data)
        
        return data
    
    def _export_csv(self, data: List[Dict], config: Dict) -> Dict[str, Any]:
        """Export data as CSV"""
        # Configuration options
        delimiter = config.get('delimiter', ',')
        include_headers = config.get('include_headers', True)
        encoding = config.get('encoding', 'utf-8')
        flatten = config.get('flatten', True)
        
        # Flatten nested data if needed
        if flatten:
            data = [self._flatten_dict(item) for item in data]
        
        # Get all unique keys for headers
        all_keys = set()
        for item in data:
            all_keys.update(item.keys())
        headers = sorted(list(all_keys))
        
        # Create CSV file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{timestamp}.csv"
        file_path = self.export_dir / filename
        
        with open(file_path, 'w', newline='', encoding=encoding) as csvfile:
            writer = csv.DictWriter(
                csvfile,
                fieldnames=headers,
                delimiter=delimiter,
                restval=''
            )
            
            if include_headers:
                writer.writeheader()
            
            for row in data:
                writer.writerow(row)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        return {
            'file_path': str(file_path),
            'file_size': file_size,
            'row_count': len(data),
            'headers': headers
        }
    
    def _export_json(self, data: List[Dict], config: Dict) -> Dict[str, Any]:
        """Export data as JSON"""
        # Configuration options
        indent = config.get('indent', 2)
        ensure_ascii = config.get('ensure_ascii', False)
        sort_keys = config.get('sort_keys', True)
        compact = config.get('compact', False)
        
        # Create JSON file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{timestamp}.json"
        file_path = self.export_dir / filename
        
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            if compact:
                json.dump(data, jsonfile, ensure_ascii=ensure_ascii, sort_keys=sort_keys)
            else:
                json.dump(
                    data,
                    jsonfile,
                    indent=indent,
                    ensure_ascii=ensure_ascii,
                    sort_keys=sort_keys
                )
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        return {
            'file_path': str(file_path),
            'file_size': file_size,
            'row_count': len(data)
        }
    
    def _export_excel(self, data: List[Dict], config: Dict) -> Dict[str, Any]:
        """Export data as Excel file"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return {
                'error': 'openpyxl not installed. Run: pip install openpyxl'
            }
        
        # Configuration options
        sheet_name = config.get('sheet_name', 'Data')
        include_headers = config.get('include_headers', True)
        auto_filter = config.get('auto_filter', True)
        freeze_panes = config.get('freeze_panes', True)
        flatten = config.get('flatten', True)
        
        # Flatten nested data if needed
        if flatten:
            data = [self._flatten_dict(item) for item in data]
        
        # Get all unique keys for headers
        all_keys = set()
        for item in data:
            all_keys.update(item.keys())
        headers = sorted(list(all_keys))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add headers with formatting
        if include_headers:
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="CCCCCC",
                end_color="CCCCCC",
                fill_type="solid"
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
        
        # Add data
        start_row = 2 if include_headers else 1
        for row_idx, item in enumerate(data, start_row):
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header, '')
                # Handle different data types
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto filter
        if auto_filter and include_headers:
            ws.auto_filter.ref = ws.dimensions
        
        # Freeze top row
        if freeze_panes and include_headers:
            ws.freeze_panes = 'A2'
        
        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{timestamp}.xlsx"
        file_path = self.export_dir / filename
        wb.save(file_path)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        return {
            'file_path': str(file_path),
            'file_size': file_size,
            'row_count': len(data),
            'headers': headers
        }
    
    def _export_parquet(self, data: List[Dict], config: Dict) -> Dict[str, Any]:
        """Export data as Parquet file"""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
            import pandas as pd
        except ImportError:
            return {
                'error': 'pyarrow and pandas not installed. Run: pip install pyarrow pandas'
            }
        
        # Configuration options
        compression = config.get('compression', 'snappy')
        flatten = config.get('flatten', True)
        
        # Flatten nested data if needed
        if flatten:
            data = [self._flatten_dict(item) for item in data]
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Create Parquet file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{timestamp}.parquet"
        file_path = self.export_dir / filename
        
        # Write Parquet file
        df.to_parquet(
            file_path,
            compression=compression,
            engine='pyarrow'
        )
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        return {
            'file_path': str(file_path),
            'file_size': file_size,
            'row_count': len(data),
            'columns': list(df.columns)
        }
    
    def _export_xml(self, data: List[Dict], config: Dict) -> Dict[str, Any]:
        """Export data as XML"""
        import xml.etree.ElementTree as ET
        from xml.dom import minidom
        
        # Configuration options
        root_element = config.get('root_element', 'data')
        item_element = config.get('item_element', 'item')
        pretty_print = config.get('pretty_print', True)
        encoding = config.get('encoding', 'utf-8')
        
        # Create root element
        root = ET.Element(root_element)
        
        # Add data items
        for item in data:
            item_elem = ET.SubElement(root, item_element)
            self._dict_to_xml(item, item_elem)
        
        # Convert to string
        if pretty_print:
            xml_str = minidom.parseString(
                ET.tostring(root, encoding='unicode')
            ).toprettyxml(indent="  ")
        else:
            xml_str = ET.tostring(root, encoding='unicode')
        
        # Create XML file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{timestamp}.xml"
        file_path = self.export_dir / filename
        
        with open(file_path, 'w', encoding=encoding) as xmlfile:
            xmlfile.write(xml_str)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        return {
            'file_path': str(file_path),
            'file_size': file_size,
            'row_count': len(data)
        }
    
    def _dict_to_xml(self, data: Dict, parent: ET.Element) -> None:
        """Convert dictionary to XML elements"""
        for key, value in data.items():
            # Sanitize key for XML element name
            safe_key = key.replace(' ', '_').replace('-', '_')
            safe_key = ''.join(c for c in safe_key if c.isalnum() or c == '_')
            
            if isinstance(value, dict):
                child = ET.SubElement(parent, safe_key)
                self._dict_to_xml(value, child)
            elif isinstance(value, list):
                for item in value:
                    child = ET.SubElement(parent, safe_key)
                    if isinstance(item, dict):
                        self._dict_to_xml(item, child)
                    else:
                        child.text = str(item)
            else:
                child = ET.SubElement(parent, safe_key)
                child.text = str(value) if value is not None else ''
    
    def _flatten_dict(
        self,
        data: Dict,
        parent_key: str = '',
        sep: str = '_'
    ) -> Dict:
        """Flatten nested dictionary"""
        items = []
        for key, value in data.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else key
            if isinstance(value, dict):
                items.extend(
                    self._flatten_dict(value, new_key, sep=sep).items()
                )
            elif isinstance(value, list):
                # Convert list to string for CSV/Excel
                items.append((new_key, json.dumps(value)))
            else:
                items.append((new_key, value))
        return dict(items)
    
    def stream_export(
        self,
        task_id: str,
        format: str = 'json',
        chunk_size: int = 1000
    ):
        """
        Stream export for large datasets
        
        Args:
            task_id: Task ID
            format: Export format
            chunk_size: Number of records per chunk
        
        Yields:
            Data chunks in specified format
        """
        executions = TaskExecution.objects.filter(
            task_id=task_id,
            status='success'
        ).iterator(chunk_size=chunk_size)
        
        if format == 'json':
            yield '['
            first = True
            for execution in executions:
                if execution.scraped_data:
                    if not first:
                        yield ','
                    yield json.dumps(execution.scraped_data)
                    first = False
            yield ']'
        
        elif format == 'csv':
            # Stream CSV
            output = io.StringIO()
            writer = None
            
            for execution in executions:
                if execution.scraped_data:
                    data = execution.scraped_data
                    if isinstance(data, list):
                        for item in data:
                            if not writer:
                                # Initialize writer with headers
                                flat_item = self._flatten_dict(item)
                                writer = csv.DictWriter(
                                    output,
                                    fieldnames=flat_item.keys()
                                )
                                writer.writeheader()
                                yield output.getvalue()
                                output.truncate(0)
                                output.seek(0)
                            
                            writer.writerow(self._flatten_dict(item))
                            yield output.getvalue()
                            output.truncate(0)
                            output.seek(0)